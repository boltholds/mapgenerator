#utf-8

import sys
#excel parse
import openpyxl
import re
import json
from collections import deque

import logging


logger = logging.getLogger(__name__)

hat = '''#utf-8
module0	= [] #default
module1	= []
module2	= []
module3	= []

##############################################################
delay = 100 # sound delay ####################################
hz = 50 # max value in hz ####################################
bar = 16 # max value in bar ##################################
colorCodingMode = 1 # how many colors one byte encodes #######
minBrightnessRed = 128 # for colorCodingMode 2 or 3 ########## in developing
minBrightnessGreen = 128 # for colorCodingMode 2 or 3 ######## in developing
minBrightnessBlue = 128 # for colorCodingMode 2 or 3 ######### in developing
##############################################################
'''
names = {"KLall" :"клапанов", "JLall":"отсекателей","ELall":"светильников","mall":"насосов"}

file_formats = {"Python":"*.py","JSON":"*.json"}


def file_cache_error(func):
	'''Decorator for error if file not found'''
	def _wrapper(*args, **kwargs):
		try:
			result = func(*args, **kwargs)
		except FileNotFoundError:
			logger.error(f"File {args[0]} or directory not found !")
		return result	
	return _wrapper


#read columns by adres
@file_cache_error
def parse_xls(file_xls):
	'''Parse name of counter and his adress in Excel table'''
	#offset_transmiter = 512
	spliters = r"[,-]"
	
	def init_row(worksheet):
		'''Find title for coordinational in sheet'''
		title = "ARK"
		for row in range(1, worksheet.max_row):
			rowvalue = worksheet[f'B{row}'].value
			if rowvalue:
				if rowvalue[:3] == title:
					return row
		logger.error("No named title {title} in file!")
					
	def get_activesh(wookbook):
		'''Find non void sheet in workbook Excel'''
		for sheet in range(0,5):
			wookbook.active = sheet
			if wookbook.active.max_row > 4:
				return sheet
		logger.error("All sheets void in file!")
				
				
	def add_counter(counters,name,adress):
		'''Special add key in dict funkt with assert for doublicate counter name in dict, that's foolproof'''
		try:
			if counters.get(name) is None:
				counters[name] = adress
			else:
				raise (AssertionError)
		except AssertionError:
			logger.warning(f" Counture '{name}' = [ {adress} ] has a duplicate! Was remove")
			

	try:
		if file_xls.split(".")[1] == "xlsx":
			wookbook = openpyxl.load_workbook(file_xls, data_only=True)
			
			#found start row of adress name
		else:
			try:
				from xls2xlsx import XLS2XLSX
				w2x = XLS2XLSX(file_xls)
				wookbook = w2x.to_xlsx()
			except ImportError:
				logger.error(f"Install XLS2XLSX lib for support format *.xlsx" )	

		wookbook.active = get_activesh(wookbook)
		worksheet = wookbook.active
		
	except TypeError:
		logger.error(f" Not support type for parsing {type(file_xls)}" )
	
	else:
		start_row = init_row(worksheet)
		conturs = dict()
		for row in range(start_row, worksheet.max_row):
			iscontur = worksheet[f'C{row}'].value

			adress = worksheet[f'E{row}'].value
			if iscontur not in ('',' ',None):
					#List counturs exp EL1-EL2 or EL1,EL2...
					if re.findall(spliters,iscontur):
						polycounturs = re.split(spliters,iscontur)
						for iscontur in polycounturs:
							add_counter(conturs,iscontur,adress)
					else:
							add_counter(conturs,iscontur,adress)		
		return conturs

def print_map(map_dict):
	'''Print the map dict'''
	for name in map_dict:
		logger.info(f" Countur {name} on adress {map_dict[name]} \n")

#parse the hierarchy counturs on map to the JSON(dict in dict)
def wide_map(premap_dicts):
	'''Groups counturs by name and suffix GlobalGroupxGroupxSubgroupxElement 
exp EL1x1x1
its ELall - GlobalGroup
EL1- Group
EL1x1 - Subgroup
ELx1x1x1 - element

Element it's key of value addresof counter in dict
'''

	#NAME is ELx,mx,JLx
	names = r'([mMKLJ]|[EKJ][Ll])[1-9]'
	#it's xNxN exp x1 or x2x1
	sufix_pattern = r'[.xX][1-9]\d*'
	#NAMExNxN,NAMExN
	match =  names+sufix_pattern
	
	global_group = dict()#result
	try:
		for name in premap_dicts:
			res = re.fullmatch(match,name)
			suf = re.findall(sufix_pattern,name)
			nm = re.findall(names,name)
			if nm:
				glob = f"{nm[0]}" # EL1 -> EL, KL1 -> KL
				group = suf[0][1:] if len(suf) else '' #x1 -> 1
				subgroup = suf[1][1:] if len(suf)>1 else ''
				group = group.replace('.','')#if name was exp m1.2.3 
				subgroup = subgroup.replace('.','')
				main =  name[len(nm[0])].replace('.','')
				key=glob+'all'
				if global_group.get(key) is None:
					global_group[key] = dict()
				counter_name = glob+main
				if global_group[key].get(counter_name) is None:
					global_group[key][counter_name] = dict()

				if(group != ''):
					subcounter_name = counter_name+'x'+group
					if	(global_group[key][counter_name].get(subcounter_name) is None):
						global_group[key][counter_name][subcounter_name] =dict()
					if subgroup != '':
						element = subcounter_name+'x'+subgroup
						global_group[key][counter_name][subcounter_name][element] = premap_dicts[name]
					else:
						global_group[key][counter_name][subcounter_name] = premap_dicts[name]
				else:
					global_group[key][counter_name] = premap_dicts[name]
		
				#several logging best thousand comment ;)
				logger.debug(f" Is {key[0:2]}{main}")
				
				logger.debug(f" group {suf[0]} " if len(suf) else '')
				logger.debug(f" subgroup {suf[1]} " if len(suf)>1 else '')
				logger.debug("\n")


	except TypeError:
		logger.error(f" Instead of {type(global_group)} passed {type(premap_dicts)}!")
		
	return global_group

			
#dump file	
@file_cache_error	
def out_json(file_name,main_dir):
	'''Dump dict to JSON file'''
	with open(file_name,"w+") as out:
		js =json.dumps(main_dir, indent=4)
		out.write(js)


#write to map.py
@file_cache_error
def outjob(filename,map_dir):
	'''Out of *.format file'''

	#splitter from countur group	
	def hat_countr(name,fil):
		'''Insert into *.py file horizontal splter that different counturs group'''
		wide = 45
		final =  f" Контур {name} ".upper()
		fil.write("\n")
		fil.write("#"*10)
		fil.write(final)
		fil.write("#"*(wide-len(final)-10))
		fil.write("\n"*2)

		
	def rec(outfile,mapy,keys,group,lvl,old_lvl):
		'''Func for recursive walk on tree heararhy
		if we down to deep, then add eque name node
		if we up from deep ,the pop eaue name node'''
		old_lvl = lvl
		lvl+=1
		if (names.get(keys) != None)&(lvl<=0):
			#if on head then print the hat :)
			hat_countr(names.get(keys),outfile)
		for key,value in mapy.items(): 
		    #if it's terminal
		    if type(value) == int:
		        outfile.write(f"{key} = [ {value} ]\n")
			#if it vertex
		    elif type(value) == dict:
		        
		        group.append(f"{key} = " +" + ".join(list(value.keys()))+"\n ") 
		        old_lvl = lvl
		        lvl+=rec(outfile,value,key,group,lvl,old_lvl)
		#print countur hierarhy
		outfile.write("\n")
		outfile.write(group.pop()) 
		outfile.write("\n")  
		return -1 #up level on JSON
		

	if map_dir:
		with open(filename,"w+") as out:
			out.write(hat)
			out.write("\n"*2)
			subgroup_name = ''
			g = deque()
			g.append('')
			rec(out,map_dir,'mall',g,-1,0)
	else:
		logger.error(f"Passed {map_dir} is void!")


if __name__ == '__main__':
	logger.setLevel(logging.INFO)
	import argparse
	parser = argparse.ArgumentParser("Translation Exlel adress table to map.py file")
	parser.add_argument('table',type=str,help='path to xlsx file')
	parser.add_argument('output', type=str, help='Output format file, default map.py')
	args  = parser.parse_args()
	der = parse_xls(args.table)
	file_type = args.output.split(".")[1]
	der = wide_map(der)
	if  file_type== "py":
		outjob(args.output,der)
	elif file_type == "json":
		out_json(args.output,der)
	else:
		print(f"Unknown file format *.{file_type}! Please try {','.join(list(file_formats.values()))} ")

